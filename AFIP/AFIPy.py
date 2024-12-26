import shutil
import os
import time
import pandas as pd
from typing import Any
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Alignment
from selenium import webdriver
from selenium.webdriver.remote.webdriver import WebDriver
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

# -------------------------------------------------------------------------------------------------
# Listas.
# -------------------------------------------------------------------------------------------------

Tipos_de_Comprobantes = {'Factura C': 2,
                         'Nota de Débito C': 3,
                         'Nota de Crédito C': 4,
                         'Recibo C': 5,
                         'Factura de Crédito Electrónica MiPyMEs (FCE) C': 120,
                         'Nota de Débito Electrónica MiPyMEs (FCE) C': 121,
                         'Nota de Crédito Electrónica MiPyMEs (FCE) C': 122}

Tipos_de_Conceptos = {'Productos': 1,
                      'Servicios': 2,
                      'Productos y Servicios': 3}

Tipos_Condicion_IVA = {"IVA Responsable Inscripto": "1",
                       "IVA Sujeto Exento": "4",
                       "Consumidor Final": "5",
                       "Responsable Monotributo": "6",
                       "Sujeto No Categorizado": "7",
                       "Proveedor del Exterior": "8",
                       "Cliente del Exterior": "9",
                       "IVA Liberado - Ley Nº 19.640": "10",
                       "Monotributista Social": "13",
                       "IVA No Alcanzado": "15",
                       "Monotributista Trabajador Independiente Promovido": "16"}


# -------------------------------------------------------------------------------------------------
# Funciones.
# -------------------------------------------------------------------------------------------------

# Etapa 1. Payway.

def Download_CSV_From_Payway(Driver: WebDriver, Email: str, Password: str) -> WebDriver:

    """
    Automates the login process and downloads a CSV file from the Payway website.

    Parameters:
        Driver: The Selenium WebDriver instance to control the browser.
        Email: The email address used for login.
        Password: The password used for login.

    Returns:
        The WebDriver instance after downloading the CSV file.

    Example:
        >>> Driver = SomeWebDriverInstance()
        >>> Download_CSV_From_Payway(Driver, "test@example.com", "password123")

    """

    Login_URL = 'https://mi.payway.com.ar/ms/ui-login/login'

    # Open the login page.
    Driver.get(Login_URL)

    # Wait until the "Email" input field is present.
    WebDriverWait(Driver, 30).until(
        EC.presence_of_element_located((By.NAME, "email"))
    )

    # Locate the email input field by its name.
    Email_Field = Driver.find_element(By.NAME, "email")

    # Enter the email into the input field.
    Email_Field.send_keys(Email)

    # Locate the password input field by its name.
    Password_Field = Driver.find_element(By.NAME, "pwd")

    # Enter the password into the input field.
    Password_Field.send_keys(Password)

    # Press Enter to submit the login form.
    Password_Field.send_keys(Keys.RETURN)

    # Wait for 15 seconds to ensure the page loads fully.
    time.sleep(15)

    Movements_URL = 'https://mi.payway.com.ar/movimientos/en-linea'

    # Open the movements page.
    Driver.get(Movements_URL)

    # Wait until the "Download CSV" button is present.
    Download_Button = WebDriverWait(Driver, 30).until(
        EC.presence_of_element_located((By.CLASS_NAME, "sc-crHmcD.dMPykX"))
    )

    # Click the "Download CSV" button.
    Download_Button.click()

    # Wait until the "Comma-separated CSV" button is present.
    Comma_CSV_Button = WebDriverWait(Driver, 30).until(
        EC.presence_of_element_located(
            (By.XPATH, "//button[@class='sc-gJbFto fMZkYs']//span[text()='Comas']")
        )
    )

    # Click the "Comma-separated CSV" button.
    Comma_CSV_Button.click()

    return Driver


# Etapa 2. Procesamiento del archivo CSV.

def Apply_Excel_Formatting(File_Path: str) -> None:

    """
    Applies specific formatting to an Excel file using OpenPyXL.

    Parameters:
        File_Path: The path to the Excel file to format.

    Returns:
        None

    Example:
        >>> Apply_Excel_Formatting('output.xlsx')
    """

    # Check if the file exists.
    if not os.path.isfile(File_Path):
        raise FileNotFoundError(f"The file '{File_Path}' does not exist.")
    
    # Load the workbook and select the active sheet.
    Workbook = load_workbook(File_Path)

    # Verify that the workbook has sheets.
    if not Workbook.sheetnames:
        raise ValueError("No sheets found in the workbook.")
    Sheet = Workbook.active

    # Verify that the sheet has content.
    if Sheet.max_row == 0 or Sheet.max_column == 0: # type: ignore
        raise ValueError("The active sheet is empty and has no content to format.")

    # Set column widths.
    Sheet.column_dimensions['A'].width = 15 # type: ignore
    Sheet.column_dimensions['B'].width = 40 # type: ignore
    Sheet.column_dimensions['C'].width = 15 # type: ignore

    # Apply general style with no decimals for column C ("Precio").
    General_Style = NamedStyle(name="general")
    General_Style.number_format = '0'

    if 'C' in Sheet.column_dimensions: # type: ignore
        for Cell in Sheet['C']: # type: ignore
            Cell.style = General_Style

    # Center align all cells.
    Center_Alignment = Alignment(horizontal='center', vertical='center')

    for Column in Sheet.iter_cols():  # type: ignore
        # Use iter_cols to iterate reliably over columns
        for Cell in Column:
            Cell.alignment = Center_Alignment

    # Save the formatted workbook.
    Workbook.save(File_Path)

def Move_And_Rename_File(Original_Path: str, Original_Name: str, New_Path: str, New_Name: str) -> None:

    """
    Moves a file from its original location to a new specified location 
    and renames it in the process.

    Parameters:
        Original_Path: The path of the directory containing the original file.
        Original_Name: The name of the file to be moved.
        New_Path: The path of the new directory where the file will be moved.
        New_Name: The new name to assign to the file after moving.

    Returns:
        None

    Example:
        >>> Move_And_Rename_File(
        ...     "C:/original", 
        ...     "file.txt", 
        ...     "C:/new/location", 
        ...     "new_file_name.txt"
        ... )

    """

    # Combine the original path and name to form the full original file path.
    Full_Original_Path = os.path.join(Original_Path, Original_Name)

    # Combine the new path and new name to form the full destination file path.
    Full_New_Path = os.path.join(New_Path, New_Name)

    # Use shutil to move and rename the file to the new location.
    shutil.move(Full_Original_Path, Full_New_Path)

def Process_And_Save_Dataframe(Dataframe: pd.DataFrame, Output_Path: str):

    """
    Processes a Pandas DataFrame, applies formatting, and saves it as an Excel file.

    Parameters:
        Dataframe: The DataFrame to process and save.
        Output_Path: The file path to save the Excel file.

    Returns:
        None

    Example:
        >>> df = pd.DataFrame({'Fecha': ['02/04/2024'], 'Precio': [100]})
        >>> Process_And_Save_Dataframe(df, 'output.xlsx')

    """

    # Columns.
    Column_Order = ['Fecha', 'Descripción', 'Precio']

    # Create a new DataFrame.
    df = pd.DataFrame(columns = Column_Order)

    # Add the 'Descripción' and 'Precio' columns from the original DataFrame.
    df['Precio'] = Dataframe['MONTO_BRUTO']
    df['Fecha'] = Dataframe['FECHA']

    # Dataframe with prices.
    System = 'J:/My Drive/Forraje/Exportar.xls'
    df_System = pd.read_excel(System)

    # Internal function to assign a description based on the closest price.
    def Assign_Description(DataFrame: pd.DataFrame, Price: float) -> str:

        """
        Assigns a description from the dataframe based on the closest price.

        This function calculates the absolute difference between the provided price 
        and the prices in the 'Precio' column of the dataframe. It then selects the 
        description from the 'Descripción' column corresponding to the closest price.

        Parameters:
        - DataFrame (pd.DataFrame): A pandas DataFrame that must contain the columns 'Precio' and 'Descripción'.
        - Price (float): The target price to find the closest match in the dataframe.

        Returns:
        - str: The description associated with the closest price found in the dataframe.

        """

        # Finds the index of the closest price.
        Closest_Index = (DataFrame['Precio'] - Price).abs().idxmin()  

        # Returns the corresponding description.
        return str(DataFrame.loc[Closest_Index, 'Descripción'])

    # Apply the function to the DataFrame.
    df['Descripción'] = df.apply(lambda Row: Assign_Description(df_System, Row['Precio']), axis=1)

    # Convert 'Fecha' column to datetime and sort by date.
    df['Fecha'] = pd.to_datetime(df['Fecha'], format='%d/%m/%Y')
    df = df.sort_values(by='Fecha')

    # Format 'Fecha' column back to 'dd/mm/yyyy' format.
    df['Fecha'] = df['Fecha'].dt.strftime('%d/%m/%Y')

    # Reset the index.
    df = df.reset_index(drop=True)

    # Save df to Excel.
    df.to_excel(Output_Path, index=False)

    # Apply additional formatting using OpenPyXL.
    Apply_Excel_Formatting(Output_Path)

def Split_Rows_By_Threshold(Data_Frame: pd.DataFrame, Column: str, Threshold: int) -> pd.DataFrame:

    """
    Splits rows in the dataframe based on a threshold value in a specified column.
    If a value exceeds the threshold, it divides the value by half of the threshold 
    and creates new rows with that value.

    Parameters:
        Data_Frame (pd.DataFrame): The input dataframe.
        Column (str): The column name where the threshold will be applied.
        Threshold (int): The threshold value for splitting.

    Returns:
        pd.DataFrame: A new dataframe with the modified rows.

    """

    # Create an empty list to store the new rows
    New_Rows = []

    # Loop through each row in the dataframe
    for _, Row in Data_Frame.iterrows():
        Value = Row[Column]
        # If the value exceeds the threshold
        if Value > Threshold:
            # Calculate how many rows to create
            Num_Rows = Value // (Threshold // 2)
            Remainder = Value % (Threshold // 2)

            # Create new rows with the split values
            for _ in range(Num_Rows):
                New_Row = Row.copy()  # Create a copy of the row
                New_Row[Column] = Threshold // 2
                New_Rows.append(New_Row)

            # Add the remainder as an extra row if any
            if Remainder > 0:
                New_Row = Row.copy()
                New_Row[Column] = Remainder
                New_Rows.append(New_Row)
        else:
            # If the value is below the threshold, keep the original row
            New_Rows.append(Row)

    # Return the new dataframe
    return pd.DataFrame(New_Rows)

# Etapa 3. AFIP.

def Initialize_Chrome_Driver() -> webdriver.Chrome:

    """
    Initializes a Chrome WebDriver instance with specific options.

    Returns:
        A Chrome WebDriver instance.

    Example:
        >>> driver = Initialize_Chrome_Driver()
    """

    Options = webdriver.ChromeOptions()
    Options.add_argument('--ignore-certificate-errors')
    Options.add_argument('--ignore-ssl-errors')

    Driver = webdriver.Chrome(options=Options)
    return Driver

def Wait_For_Download(File_Path: str, Timeout: int = 30, Interval: int = 2) -> None:

    """
    Waits for a file to be downloaded to the specified path.

    Parameters:
        File_Path: The path to the expected downloaded file.
        Timeout: Maximum waiting time in seconds.
        Interval: Time between file existence checks in seconds.

    Raises:
        TimeoutError: If the file is not found within the timeout.

    Example:
        >>> Wait_For_Download("/path/to/file.csv")
    """

    Start_Time = time.time()
    while not os.path.exists(File_Path):
        if time.time() - Start_Time > Timeout:
            raise TimeoutError("Download timeout exceeded.")
        time.sleep(Interval)

def Login_To_AFIP(Driver: webdriver.Chrome, Username: str, Password: str, Company: str) -> webdriver.Chrome:

    """
    Logs into the AFIP platform.

    Parameters:
        Driver: The Selenium WebDriver instance.
        Username: The AFIP username.
        Password: The AFIP password.
        Company: The company to select.

    Returns:
        The WebDriver instance after login.

    Example:
        >>> Login_To_AFIP(driver, "user@example.com", "password123", "My Company")

    """

    Login_URL = 'https://auth.afip.gob.ar/contribuyente_/login.xhtml'
    Driver.get(Login_URL)

    WebDriverWait(Driver, 10).until(
        EC.presence_of_element_located((By.ID, 'F1:username'))
    )

    Username_Field = Driver.find_element(By.ID, 'F1:username')
    Username_Field.send_keys(Username)
    Username_Field.send_keys(Keys.RETURN)

    WebDriverWait(Driver, 10).until(
        EC.presence_of_element_located((By.ID, 'F1:password'))
    )

    Password_Field = Driver.find_element(By.ID, 'F1:password')
    Password_Field.send_keys(Password)
    Password_Field.send_keys(Keys.RETURN)

    # Navigate to the invoice section.
    WebDriverWait(Driver, 10).until(EC.presence_of_element_located(
        (By.CSS_SELECTOR, 'h3.roboto-font.regular.p-y-0.m-y-0.h4')))
    Invoices_Link = Driver.find_element(By.CSS_SELECTOR, 'h3.roboto-font.regular.p-y-0.m-y-0.h4')
    Invoices_Link.click()

    WebDriverWait(Driver, 10).until(EC.number_of_windows_to_be(2))
    Windows = Driver.window_handles
    Driver.switch_to.window(Windows[1])

    Company_XPath = f"//input[@value='{Company}']"  
    Company_Button = Driver.find_element(By.XPATH, Company_XPath)
    Company_Button.click()

    return Driver

def Generate_Invoice(Driver: webdriver.Chrome, Sales_Point: str, Invoice_Type: int, 
                     Date: str, Concept_Type: int, VAT_Type: int, Payment_Type: str,
                     Description: str, Price: float) -> webdriver.Chrome:

    """
    Generates an invoice using the AFIP online system.

    Parameters:
        Driver (webdriver.Chrome): Selenium WebDriver instance.
        Sales_Point (str): Sales point identifier.
        Invoice_Type (int): Type of invoice to be generated.
        Date (str): Invoice date in "YYYY-MM-DD" format.
        Concept_Type (int): Concept type for the invoice.
        VAT_Type (int): VAT condition of the recipient.
        Payment_Type (str): Type of payment (e.g., 'Contado').
        Description (str): Description of the product/service.
        Price (float): Price of the product/service.

    Returns:
        webdriver.Chrome: The WebDriver instance after completing the process.

    Example:
        >>> Generate_Invoice(Driver, "0001", 2, "2024-12-26", 2, 1, "Contado", "Service", 1000.0)

    """

    # Step 1: Generate Invoice
    WebDriverWait(Driver, 10).until(EC.presence_of_element_located(
        (By.XPATH, "//span[@class='ui-button-text' and text()='Generar Comprobantes']")))
    Generate_Button = Driver.find_element(By.XPATH, 
        "//span[@class='ui-button-text' and text()='Generar Comprobantes']")
    Generate_Button.click()

    # Step 2: Select Sales Point and Invoice Type
    WebDriverWait(Driver, 10).until(EC.presence_of_element_located((By.ID, "puntodeventa")))
    Sales_Point_List = Driver.find_element(By.ID, "puntodeventa")
    Sales_Point_Option = Sales_Point_List.find_element(By.XPATH, f"//option[contains(text(), '{Sales_Point}')]")
    Sales_Point_Option.click()

    WebDriverWait(Driver, 10).until(EC.presence_of_element_located((By.ID, "universocomprobante")))
    Invoice_Type_List = Driver.find_element(By.ID, "universocomprobante")
    Select(Invoice_Type_List).select_by_value(str(Invoice_Type))

    Continue_Button = Driver.find_element(By.XPATH, "//input[@value='Continuar >']")
    Continue_Button.click()

    # Step 3: Enter Date and Concept
    WebDriverWait(Driver, 10).until(EC.presence_of_element_located((By.ID, "fc")))
    Date_Field = Driver.find_element(By.ID, "fc")
    Date_Field.clear()
    Date_Field.send_keys(Date)

    Concept_List = Driver.find_element(By.ID, "idconcepto")
    Select(Concept_List).select_by_value(str(Concept_Type))
    Continue_Button = Driver.find_element(By.XPATH, "//input[@value='Continuar >']")
    Continue_Button.click()

    # Step 4: Select VAT Condition and Payment Type
    WebDriverWait(Driver, 10).until(EC.presence_of_element_located((By.ID, "idivareceptor")))
    VAT_List = Driver.find_element(By.ID, "idivareceptor")
    Select(VAT_List).select_by_value(str(VAT_Type))

    if Payment_Type.lower() == 'contado':
        Payment_Button = Driver.find_element(By.ID, "formadepago1")
        Payment_Button.click()

    Continue_Button = Driver.find_element(By.XPATH, "//input[@value='Continuar >']")
    Continue_Button.click()

    # Step 5: Enter Product Description and Price
    WebDriverWait(Driver, 10).until(EC.presence_of_element_located((By.ID, "detalle_descripcion1")))
    Description_Field = Driver.find_element(By.ID, "detalle_descripcion1")
    Description_Field.clear()
    Description_Field.send_keys(Description)

    Price_Field = Driver.find_element(By.ID, "detalle_precio1")
    Price_Field.clear()
    Price_Field.send_keys(str(Price))

    Continue_Button = Driver.find_element(By.XPATH, "//input[@value='Continuar >']")
    Continue_Button.click()

    # Step 6: Confirm
    WebDriverWait(Driver, 10).until(EC.presence_of_element_located((By.XPATH, "//input[@value='Confirmar Datos...']")))
    Confirm_Button = Driver.find_element(By.XPATH, "//input[@value='Confirmar Datos...']")
    Confirm_Button.click()

    Alert = WebDriverWait(Driver, 10).until(EC.alert_is_present())
    Alert.accept()

    Menu_Button = Driver.find_element(By.XPATH, "//input[@value='Menú Principal']")
    Menu_Button.click()

    return Driver


# Etapa 4. Cierre.

def Close_All_Chrome_Tabs(Driver: WebDriver) -> None:

    """
    Closes all Chrome tabs opened by the given WebDriver instance.

    Parameters:
        Driver: The Selenium WebDriver instance controlling the browser.

    Returns:
        None

    Example:
        >>> Close_All_Chrome_Tabs(driver)
    """

    # Close all tabs by quitting the WebDriver session.
    Driver.quit()