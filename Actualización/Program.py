# ------------------------------------------------------------------------------------------------
# Paquetes.
# ------------------------------------------------------------------------------------------------

import pandas as pd
import numpy as np
import math 
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side, Alignment, Color
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.worksheet.worksheet import Worksheet

import sys
sys.path.append('c:/Users/tomas/Documents/Programación/Proyectos/Paquetes')
import Forpy as fg
import Excpy as ex
import Frampy as fr
import Stringpy as st


# ------------------------------------------------------------------------------------------------
# Variables.
# ------------------------------------------------------------------------------------------------

# Archivo Proveedores.
Proveedores = 'J:/My Drive/Forraje/Proveedores.xlsx'
df_Proveedores = pd.read_excel(Proveedores, sheet_name='Centro')

# Proveedores.
Lista_Proveedores = ['Benavídez', 'Gallo', 'Anea', 'Pilar', 'Tortuguitas', 'Del Viso',
                     'Chica Glade', 'San Martín', 'Hipocampus', 'Mundo Mascotas', 'Yaguar',
                     'King Clor', 'Diserquim', 'Moreno', 'Roberto', 'Viejo Gets', 'Viejo Cuero', 'Mis Nietos', 
                     'Profesional Vet', 'Premier', 'Dacam']

# Archivo Nex.
Nex = 'J:/My Drive/Forraje/Exportar.xls'
df_Nex = pd.read_excel(Nex)

# Archivo Granel.
Granel = 'J:/My Drive/Forraje/Granel.xlsx'
df_Granel = pd.read_excel(Granel)

# Archivo Fraccionados.
Fraccion = 'J:/My Drive/Forraje/Fraccionables.xls'
df_Fraccion = pd.read_excel(Fraccion)

# Ofertas y precios establecidos.
Ofertas = {}



# ------------------------------------------------------------------------------------------------
# ------------------------------------------------------------------------------------------------
# ------------------------------------------------------------------------------------------------
# PROCESAMIENTO DEL ARCHIVO PROVEEDORES.XLSX
# ------------------------------------------------------------------------------------------------
# ------------------------------------------------------------------------------------------------
# ------------------------------------------------------------------------------------------------

# ------------------------------------------------------------------------------------------------
# Formatear el texto de las columnas de proveedores.
# ------------------------------------------------------------------------------------------------

df_Proveedores = fg.Column_Provider_Processing(df_Proveedores, Lista_Proveedores)


# ------------------------------------------------------------------------------------------------
# Hallar precio mínimo y proveedor.
# ------------------------------------------------------------------------------------------------

Minimo_Sin_Cero = df_Proveedores[Lista_Proveedores].replace(0, np.nan).min(axis=1, skipna=True)
df_Proveedores['Precio'] = Minimo_Sin_Cero

df_Proveedores = fg.Find_Best_Provider(df_Proveedores, Lista_Proveedores)


# ------------------------------------------------------------------------------------------------
# Agregar columnas con los precios y proveedores más baratos.
# ------------------------------------------------------------------------------------------------

Columns = ['Codigo', 'Descripcion', 'Precio', 'Categoria', 'Proveedor']

df_Proveedores = df_Proveedores[Columns].copy()
df_Proveedores.rename(columns={'Precio': 'Costo'}, inplace=True)


# ------------------------------------------------------------------------------------------------
# Agregar columnas con Markups y tipo de Unidad.
# ------------------------------------------------------------------------------------------------

df_Proveedores['Markup %'] = df_Proveedores['Categoria'].map(lambda x: fg.Markup_And_Unity[x]['Markup %'])
df_Proveedores['Unidad'] = df_Proveedores['Categoria'].map(lambda x: fg.Markup_And_Unity[x]['Unidad'])


# ------------------------------------------------------------------------------------------------
# Filtrar df.
# ------------------------------------------------------------------------------------------------

Columns = ['Descripcion','Costo','Markup %','Proveedor','Categoria', 'Unidad']
df_Proveedores = df_Proveedores[Columns]



# ------------------------------------------------------------------------------------------------
# ------------------------------------------------------------------------------------------------
# ------------------------------------------------------------------------------------------------
# PROCESAMIENTO DEL ARCHIVO EXPORTAR.XLS, CATÁLOGO DEL SISTEMA NEXTAR
# ------------------------------------------------------------------------------------------------
# ------------------------------------------------------------------------------------------------
# ------------------------------------------------------------------------------------------------
# ------------------------------------------------------------------------------------------------

# ------------------------------------------------------------------------------------------------
# Renombrar columnas y filtrar df.
# ------------------------------------------------------------------------------------------------

df_Nex.rename(columns={'Costo Unitário': 'Costo'}, inplace=True)

df_Nex = fr.Add_Word_To_Name_Columns(df_Nex, Word = "Nex", Separator = '_')

# ------------------------------------------------------------------------------------------------
# ------------------------------------------------------------------------------------------------
# ------------------------------------------------------------------------------------------------
# DATAFRAME FINAL.
# ------------------------------------------------------------------------------------------------
# ------------------------------------------------------------------------------------------------
# ------------------------------------------------------------------------------------------------
# ------------------------------------------------------------------------------------------------

# ------------------------------------------------------------------------------------------------
# Verificar que los dos df tienen el mismo largo y componentes.
# ------------------------------------------------------------------------------------------------

Is_Same_Lenght = fr.Compare_Columns(df_Nex, 'Descripcion_Nex', df_Proveedores, 'Descripcion')

if len(Is_Same_Lenght) != 0:
    print(Is_Same_Lenght)
    raise Exception("Los dfs no tienen la misma longitud o tienen diferencias en los nombres de los productos.")

Columns = ['Codigo_Nex','Costo_Nex','Precio_Nex']
df_Nex = df_Nex[Columns]

df = pd.concat([df_Nex, df_Proveedores], axis=1)


df.rename(columns={'Codigo_Nex': 'Codigo'}, inplace=True)
df.rename(columns={'Costo': 'Costo_Prov'}, inplace=True)
df.rename(columns={'Precio': 'Precio_Nex'}, inplace=True)

Columns_New_Order = ["Codigo", "Descripcion", "Costo Final", "Precio_Nex", "Costo_Nex", "Costo_Prov", "Markup %", "Categoria", "Proveedor"]

df = df.reindex(columns=Columns_New_Order)

# ------------------------------------------------------------------------------------------------
# Calcular variables nuevas.
# ------------------------------------------------------------------------------------------------

df['Costo Final'] = df.apply(lambda row: max(row['Costo_Nex'], row['Costo_Prov']), axis=1).fillna(0)


# ------------------------------------------------------------------------------------------------
# ------------------------------------------------------------------------------------------------
# ------------------------------------------------------------------------------------------------
# CALCULAR PRECIOS A GRANEL.
# ------------------------------------------------------------------------------------------------
# ------------------------------------------------------------------------------------------------
# ------------------------------------------------------------------------------------------------
# ------------------------------------------------------------------------------------------------

df_Granel = fr.Get_Last_Number_Of_String_Column(df_Granel, 'Bolsa', 'Kilos')

df_Granel = fr.Match_And_Copy_Column_Values(df_Granel, df, 'Bolsa', 'Costo Bolsa', 'Descripcion',  'Costo Final')

df_Granel = df_Granel.replace('', 0)

df_Granel = fr.Convert_Type_Of_Columns(df_Granel, ['Kilos', 'Costo Bolsa', 'Costo Final'], "float")

df_Granel['Costo Kilo'] = df_Granel['Costo Bolsa']/df_Granel['Kilos']

df = fr.Match_And_Copy_Column_Values(df, df_Granel, 'Descripcion', 'Costo_Prov', 'Granel', 'Costo Kilo')



# ------------------------------------------------------------------------------------------------
# ------------------------------------------------------------------------------------------------
# ------------------------------------------------------------------------------------------------
# CALCULAR PRECIOS DE FRACCIONADOS.
# ------------------------------------------------------------------------------------------------
# ------------------------------------------------------------------------------------------------
# ------------------------------------------------------------------------------------------------
# ------------------------------------------------------------------------------------------------

df_Fraccion = fr.Get_Last_Number_Of_String_Column(df_Fraccion, 'Fraccionable', 'Unidades')

df_Fraccion = fr.Match_And_Copy_Column_Values(df_Fraccion, df, 'Fraccionable', 'Costo Paquete', 'Descripcion', 'Costo Final')

df_Fraccion = df_Fraccion.replace('', 0)

df_Fraccion = fr.Convert_Type_Of_Columns(df_Fraccion, ['Unidades', 'Costo Paquete'], "float")

df_Fraccion['Costo Unidad'] = df_Fraccion['Costo Paquete']/df_Fraccion['Unidades']

# ------------------------------------------------------------------------------------------------
# Casos especiales: Ivomec.
# ------------------------------------------------------------------------------------------------


